from flask import Blueprint, render_template, request, redirect, url_for, flash, session, json, current_app, Response, jsonify
from datetime import datetime, timedelta, date
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename
from ..db import get_db, calculate_accrued_days
from .. import ad_sync
import sqlite3
import os
import io
from openpyxl import Workbook
from ..utils import send_email, get_paraguay_holidays, calculate_working_days

bp = Blueprint('hr', __name__, url_prefix='/hr')

def check_hr_access(readonly=False):
    role = session.get("base_role")
    if role == "RRHH":
        return True
    if readonly and role == "Asistente RRHH":
        return True
    return False

@bp.route("/generate_periods")
def generate_periods():
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    current_year = datetime.now().year
    db = get_db()
    employees = db.execute("SELECT id, hire_date FROM employees WHERE is_active = 1;").fetchall()
    vac_type = db.execute("SELECT id FROM leave_types WHERE name = 'Vacaciones'").fetchone()
    vac_type_id = vac_type['id'] if vac_type else 1

    generated_count = 0
    for emp in employees:
        emp_id = emp["id"]
        hire_date = emp["hire_date"]
        period_exists = db.execute(
            "SELECT id FROM vacation_periods WHERE employee_id = ? AND year = ? AND leave_type_id = ?;",
            (emp_id, current_year, vac_type_id)
        ).fetchone()
        
        if period_exists is None:
            accrued_days = calculate_accrued_days(hire_date)
            db.execute(
                "INSERT INTO vacation_periods (employee_id, year, leave_type_id, total_days_accrued) VALUES (?, ?, ?, ?);",
                (emp_id, current_year, vac_type_id, accrued_days)
            )
            generated_count += 1

    db.commit()
    flash(f"Se generaron/verificaron {generated_count} nuevos periodos para el año {current_year}.", "success")
    return redirect(url_for('main.dashboard'))

@bp.route('/period/add', methods=('GET', 'POST'))
def hr_add_period():
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    db = get_db()

    if request.method == 'POST':
        employee_id = request.form.get('employee_id')
        year = request.form.get('year')
        total_days = request.form.get('total_days_accrued')
        leave_type_id = request.form.get('leave_type_id')

        if not all([employee_id, year, total_days]):
            flash("Todos los campos son obligatorios.", "danger")
            return redirect(url_for('hr.hr_add_period'))

        try:
            year = int(year)
            total_days = float(total_days)
        except (ValueError, TypeError):
            flash("El año y los días deben ser números válidos.", "danger")
            return redirect(url_for('hr.hr_add_period'))

        existing_period = db.execute(
            "SELECT id FROM vacation_periods WHERE employee_id = ? AND year = ? AND leave_type_id = ?",
            (employee_id, year, leave_type_id)
        ).fetchone()

        if existing_period:
            flash(f"Ya existe un saldo para este tipo de licencia en el año {year}.", "danger")
        else:
            db.execute(
                "INSERT INTO vacation_periods (employee_id, year, leave_type_id, total_days_accrued, days_taken) VALUES (?, ?, ?, ?, 0)",
                (employee_id, year, leave_type_id, total_days)
            )
            db.commit()
            flash(f"Saldo asignado exitosamente para el año {year}.", "success")
            return redirect(url_for('hr.hr_period_list'))

    employees = db.execute("SELECT id, full_name FROM employees WHERE is_active = 1 ORDER BY full_name").fetchall()
    leave_types = db.execute("SELECT * FROM leave_types ORDER BY name").fetchall()
    
    return render_template('hr/hr_period_form.html', form_title="Asignar Saldo de Licencia", employees=employees, leave_types=leave_types, now=datetime.now)

@bp.route('/periods', methods=('GET',))
def hr_period_list():
    if not check_hr_access(readonly=True):
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))
    
    db = get_db()
    filter_employee_ids = request.args.getlist('employee_id')

    query = """
        SELECT vp.id, vp.year, vp.total_days_accrued, vp.days_taken, vp.adjustment_comment, e.full_name, lt.name as leave_name
        FROM vacation_periods vp
        JOIN employees e ON vp.employee_id = e.id
        JOIN leave_types lt ON vp.leave_type_id = lt.id
    """
    params = []
    if filter_employee_ids:
        placeholders = ','.join(['?'] * len(filter_employee_ids))
        query += f" WHERE vp.employee_id IN ({placeholders})"
        params.extend(filter_employee_ids)
    
    query += " ORDER BY e.full_name, vp.year DESC"

    periods = db.execute(query, params).fetchall()

    # Exportar a Excel (CSV)
    if request.args.get('export') == 'true':
        wb = Workbook()
        ws = wb.active
        ws.title = "Periodos"
        ws.append(['Empleado', 'Tipo Licencia', 'Año', 'Días Otorgados', 'Días Tomados', 'Saldo', 'Comentario'])
        
        for p in periods:
            saldo = p['total_days_accrued'] - p['days_taken']
            ws.append([p['full_name'], p['leave_name'], p['year'], p['total_days_accrued'], p['days_taken'], saldo, p['adjustment_comment']])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": "attachment; filename=periodos_vacaciones.xlsx"}
        )

    employees = db.execute("SELECT id, full_name FROM employees WHERE is_active = 1 ORDER BY full_name").fetchall()

    return render_template('hr/hr_period_list.html', 
                           periods=periods, 
                           employees=employees, 
                           filters={'employee_id': filter_employee_ids})

@bp.route('/period/edit/<int:period_id>', methods=('GET', 'POST'))
def hr_edit_period(period_id):
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))
    
    db = get_db()
    period = db.execute(
        """
        SELECT vp.id, vp.year, vp.total_days_accrued, vp.days_taken, vp.adjustment_comment, e.full_name, e.id as employee_id, lt.name as leave_name, vp.leave_type_id
        FROM vacation_periods vp
        JOIN employees e ON vp.employee_id = e.id
        JOIN leave_types lt ON vp.leave_type_id = lt.id
        WHERE vp.id = ?
        """, (period_id,)
    ).fetchone()

    if not period:
        flash("Periodo no encontrado.", "danger")
        return redirect(url_for('hr.hr_period_list'))

    # Fetch all periods for this employee
    employee_periods = db.execute(
        """
        SELECT vp.id, vp.year, vp.total_days_accrued, vp.days_taken, lt.name as leave_name
        FROM vacation_periods vp
        JOIN leave_types lt ON vp.leave_type_id = lt.id
        WHERE vp.employee_id = ?
        ORDER BY vp.year DESC
        """, (period['employee_id'],)
    ).fetchall()

    # Fetch leave types for the add form
    leave_types = db.execute("SELECT * FROM leave_types ORDER BY name").fetchall()

    if request.method == 'POST':
        if request.form.get('action') == 'add_license':
            try:
                new_year = datetime.now().year
                new_leave_type = request.form.get('new_leave_type_id')

                if not new_leave_type:
                    flash("Todos los campos son obligatorios para asignar una licencia.", "danger")
                else:
                    exists = db.execute(
                        "SELECT id FROM vacation_periods WHERE employee_id = ? AND year = ? AND leave_type_id = ?",
                        (period['employee_id'], new_year, new_leave_type)
                    ).fetchone()
                    
                    if exists:
                        flash(f"Ya existe un saldo para este tipo de licencia en el año {new_year}.", "danger")
                    else:
                        # Calcular días automáticamente según el tipo de licencia
                        lt_info = db.execute("SELECT default_days, name FROM leave_types WHERE id = ?", (new_leave_type,)).fetchone()
                        
                        if lt_info['default_days'] > 0:
                            # Licencia con días fijos (ej: Maternidad, Matrimonio)
                            new_days = lt_info['default_days']
                        else:
                            # Licencia variable (ej: Vacaciones), calcular por antigüedad
                            emp_info = db.execute("SELECT hire_date FROM employees WHERE id = ?", (period['employee_id'],)).fetchone()
                            new_days = calculate_accrued_days(emp_info['hire_date'])

                        db.execute(
                            "INSERT INTO vacation_periods (employee_id, year, leave_type_id, total_days_accrued, days_taken) VALUES (?, ?, ?, ?, 0)",
                            (period['employee_id'], new_year, new_leave_type, new_days)
                        )
                        db.commit()
                        flash(f"Nueva licencia asignada exitosamente ({new_days} días).", "success")
                        return redirect(url_for('hr.hr_edit_period', period_id=period_id))
            except (ValueError, TypeError):
                flash("Error en los datos ingresados.", "danger")
        else:
            try:
                total_days = float(request.form['total_days_accrued'])
                days_taken = float(request.form['days_taken'])
                comment = request.form.get('adjustment_comment')
            except (ValueError, TypeError):
                flash("Los valores de los días deben ser números.", "danger")
                return render_template('hr/hr_period_form.html', period=period, employee_periods=employee_periods, leave_types=leave_types, now=datetime.now)

            if not comment:
                flash("Es obligatorio añadir un comentario justificando la modificación.", "danger")
                return render_template('hr/hr_period_form.html', period=period, employee_periods=employee_periods, leave_types=leave_types, now=datetime.now)

            db.execute(
                "UPDATE vacation_periods SET total_days_accrued = ?, days_taken = ?, adjustment_comment = ? WHERE id = ?",
                (total_days, days_taken, comment, period_id)
            )
            db.commit()
            flash(f"Periodo de {period['full_name']} para el año {period['year']} actualizado exitosamente.", "success")
            return redirect(url_for('hr.hr_period_list'))

    return render_template('hr/hr_period_form.html', period=period, employee_periods=employee_periods, leave_types=leave_types, now=datetime.now)

@bp.route('/email_config', methods=['GET', 'POST'])
def hr_email_config():
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))
    
    db = get_db()
    if request.method == 'POST':
        db.execute("DELETE FROM email_config")
        db.execute(
            """
            INSERT INTO email_config (MAIL_SERVER, MAIL_PORT, MAIL_USE_TLS, MAIL_USE_SSL, MAIL_USERNAME, MAIL_PASSWORD, MAIL_DEFAULT_SENDER)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                request.form.get('server'),
                int(request.form.get('port', 587)),
                'use_tls' in request.form,
                'use_ssl' in request.form,
                request.form.get('username'),
                request.form.get('password', '').replace(' ', ''), # Eliminar espacios comunes en App Passwords
                request.form.get('sender')
            )
        )
        db.commit()
        flash("Configuración de correo guardada. La aplicación se reiniciará para aplicar los cambios.", "success")
        return redirect(url_for('hr.hr_email_config'))
        
    config = db.execute("SELECT * FROM email_config ORDER BY id DESC LIMIT 1").fetchone()
    return render_template('hr/hr_email_config.html', config=config)

@bp.route('/stats')
def stats_dashboard():
    if not check_hr_access(readonly=True):
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    conn = get_db()
    current_year = str(datetime.now().year)

    # KPIs
    active_employees = conn.execute("SELECT COUNT(*) FROM employees WHERE is_active = 1").fetchone()[0]
    pending_requests = conn.execute("SELECT COUNT(*) FROM vacation_requests WHERE status = 'Pendiente' OR status = 'Aprobado por Jefe'").fetchone()[0]
    days_approved_this_year = conn.execute(
        "SELECT SUM(days_requested) FROM vacation_requests WHERE status IN ('Aprobado por RRHH', 'Activo', 'Finalizado') AND strftime('%Y', start_date) = ?", 
        (current_year,)
    ).fetchone()[0] or 0

    # Datos para el gráfico de solicitudes por mes
    requests_by_month_rows = conn.execute(
        "SELECT strftime('%m', request_date) as month, COUNT(id) as count FROM vacation_requests WHERE strftime('%Y', request_date) = ? GROUP BY month ORDER BY month",
        (current_year,)
    ).fetchall()
    
    month_labels = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    requests_per_month_data = [0] * 12
    for row in requests_by_month_rows:
        month_index = int(row['month']) - 1
        requests_per_month_data[month_index] = row['count']

    # Datos para el gráfico de días por departamento
    days_by_dept_rows = conn.execute(
        """
        SELECT e.department, SUM(vr.days_requested) as total_days 
        FROM vacation_requests vr 
        JOIN employees e ON vr.employee_id = e.id 
        WHERE vr.status IN ('Aprobado por RRHH', 'Activo', 'Finalizado') AND e.department IS NOT NULL AND e.department != ''
        GROUP BY e.department
        """
    ).fetchall()

    dept_labels = [row['department'] for row in days_by_dept_rows]
    days_per_dept_data = [row['total_days'] for row in days_by_dept_rows]

    return render_template('hr/stats_dashboard.html',
                           active_employees=active_employees,
                           pending_requests=pending_requests,
                           days_approved_this_year=days_approved_this_year,
                           month_labels=json.dumps(month_labels),
                           requests_per_month_data=json.dumps(requests_per_month_data),
                           dept_labels=json.dumps(dept_labels),
                           days_per_dept_data=json.dumps(days_per_dept_data))

@bp.route("/approvals")
def hr_approval_list():
    if not check_hr_access(readonly=True):
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    db = get_db()
    
    filter_employee_ids = request.args.getlist('employee_id')
    
    query = """
        SELECT vr.id, vr.start_date, vr.end_date, vr.days_requested, vr.replacement_name, e.full_name as employee_name, m.full_name as manager_name, lt.name as leave_name
        FROM vacation_requests vr
        JOIN employees e ON vr.employee_id = e.id
        LEFT JOIN leave_types lt ON vr.leave_type_id = lt.id
        LEFT JOIN employees m ON e.manager_id = m.id
        WHERE vr.status = 'Aprobado por Jefe'
    """
    params = []
    
    if filter_employee_ids:
        placeholders = ','.join(['?'] * len(filter_employee_ids))
        query += f" AND vr.employee_id IN ({placeholders})"
        params.extend(filter_employee_ids)
        
    query += " ORDER BY vr.request_date"
    
    hr_pending_requests = db.execute(query, params).fetchall()

    # Exportar a Excel (CSV)
    if request.args.get('export') == 'true':
        wb = Workbook()
        ws = wb.active
        ws.title = "Solicitudes Pendientes"
        ws.append(['Empleado', 'Jefe Directo', 'Tipo Licencia', 'Inicio', 'Fin', 'Días Solicitados'])
        
        for req in hr_pending_requests:
            ws.append([req['employee_name'], req['manager_name'] or 'N/A', req['leave_name'] or 'Vacaciones', req['start_date'], req['end_date'], req['days_requested']])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": "attachment; filename=solicitudes_pendientes_aprobacion.xlsx"}
        )

    employees = db.execute("SELECT id, full_name FROM employees ORDER BY full_name").fetchall()
    
    return render_template("hr/hr_approval_list.html", requests=hr_pending_requests, employees=employees, filters={'employee_id': filter_employee_ids})

@bp.route("/approve/<int:request_id>", methods=["POST"])
def hr_approve_request(request_id):
    if not check_hr_access(): # Solo RRHH (Asistente retorna False)
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    db = get_db()
    
    req = db.execute("SELECT employee_id, days_requested, leave_type_id, replacement_name FROM vacation_requests WHERE id = ? AND status = 'Aprobado por Jefe'", (request_id,)).fetchone()

    if req:
        leave_type = db.execute("SELECT requires_balance FROM leave_types WHERE id = ?", (req['leave_type_id'],)).fetchone()
        requires_balance = leave_type['requires_balance'] if leave_type else 1

        if requires_balance:
            days_to_deduct = req['days_requested']
            
            periods = db.execute(
                "SELECT id, total_days_accrued, days_taken FROM vacation_periods WHERE employee_id = ? AND total_days_accrued > days_taken ORDER BY year ASC",
                (req['employee_id'],)
            ).fetchall()

            for period in periods:
                if days_to_deduct <= 0:
                    break
                
                balance = period['total_days_accrued'] - period['days_taken']
                deduct_from_this_period = min(days_to_deduct, balance)
                
                if deduct_from_this_period > 0:
                    db.execute(
                        "UPDATE vacation_periods SET days_taken = days_taken + ? WHERE id = ?",
                        (deduct_from_this_period, period['id'])
                    )
                    days_to_deduct -= deduct_from_this_period

        db.execute(
            "UPDATE vacation_requests SET status = 'Aprobado por RRHH', hr_approval_date = ? WHERE id = ?",
            (datetime.now(), request_id)
        )
        db.commit()

        # --- NOTIFICACIÓN: A Empleado, Jefe y Reemplazo (Aprobación Final) ---
        try:
            # 1. Datos del Empleado y su Jefe
            emp_info = db.execute("SELECT email, full_name, manager_id FROM employees WHERE id = ?", (req['employee_id'],)).fetchone()
            recipients = []
            
            if emp_info and emp_info['email']:
                recipients.append(emp_info['email'])
            
            if emp_info and emp_info['manager_id']:
                mgr_info = db.execute("SELECT email FROM employees WHERE id = ?", (emp_info['manager_id'],)).fetchone()
                if mgr_info and mgr_info['email']:
                    recipients.append(mgr_info['email'])

            # 2. Datos del Reemplazo
            if req['replacement_name']:
                rep_info = db.execute("SELECT email FROM employees WHERE full_name = ?", (req['replacement_name'],)).fetchone()
                if rep_info and rep_info['email']:
                    recipients.append(rep_info['email'])

            # Obtener email de RRHH (Actor) para CC
            hr_actor_email = db.execute("SELECT email FROM employees WHERE id = ?", (session['user_id'],)).fetchone()
            cc_list = [hr_actor_email['email']] if hr_actor_email and hr_actor_email['email'] else []

            if recipients:
                subject = "Solicitud de Vacaciones Aprobada"
                body = f"Estimado/a,\n\nLa solicitud de vacaciones de {emp_info['full_name']} ha sido aprobada por RRHH y ya está activa en el sistema.\n\nReemplazo asignado: {req['replacement_name'] or 'N/A'}"
                send_email(subject, recipients, body, cc=cc_list)
        except Exception as e:
            print(f"Error enviando email al empleado: {e}")

        flash("Solicitud aprobada y saldo actualizado.", "success")
    else:
        flash("No se pudo encontrar la solicitud o ya fue procesada.", "warning")

    return redirect(url_for("hr.hr_approval_list"))

@bp.route("/reject/<int:request_id>", methods=["POST"])
def hr_reject_request(request_id):
    if not check_hr_access(): # Solo RRHH
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    db = get_db()
    
    req_to_reject = db.execute("SELECT id FROM vacation_requests WHERE id = ? AND status = 'Aprobado por Jefe'", (request_id,)).fetchone()

    if req_to_reject:
        db.execute("UPDATE vacation_requests SET status = 'Rechazado' WHERE id = ?", (request_id,))
        db.commit()

        # --- NOTIFICACIÓN: A Empleado y Jefe (Rechazo RRHH) ---
        try:
            emp_info = db.execute("SELECT e.email, e.full_name, e.manager_id FROM vacation_requests vr JOIN employees e ON vr.employee_id = e.id WHERE vr.id = ?", (request_id,)).fetchone()
            recipients = []
            
            if emp_info and emp_info['email']:
                recipients.append(emp_info['email'])
            
            if emp_info and emp_info['manager_id']:
                mgr_info = db.execute("SELECT email FROM employees WHERE id = ?", (emp_info['manager_id'],)).fetchone()
                if mgr_info and mgr_info['email']:
                    recipients.append(mgr_info['email'])

            # Obtener email de RRHH (Actor) para CC
            hr_actor_email = db.execute("SELECT email FROM employees WHERE id = ?", (session['user_id'],)).fetchone()
            cc_list = [hr_actor_email['email']] if hr_actor_email and hr_actor_email['email'] else []

            if recipients:
                subject = "Solicitud de Vacaciones Rechazada por RRHH"
                body = f"Estimado/a,\n\nLa solicitud de vacaciones de {emp_info['full_name']} ha sido rechazada por el departamento de RRHH."
                send_email(subject, recipients, body, cc=cc_list)
        except Exception as e:
            print(f"Error enviando email al empleado: {e}")

        flash("La solicitud ha sido rechazada.", "info")
    else:
        flash("No se pudo encontrar la solicitud o ya fue procesada.", "warning")

    return redirect(url_for("hr.hr_approval_list"))

@bp.route("/employees")
def hr_employee_list():
    if not check_hr_access(readonly=True):
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))
    
    db = get_db()
    employees = db.execute(
        "SELECT id, full_name FROM employees ORDER BY full_name"
    ).fetchall()
    
    filter_employee_ids = request.args.getlist('employee_id')
    
    query = """
        SELECT e.*, m.full_name as manager_name
        FROM employees e
        LEFT JOIN employees m ON e.manager_id = m.id
    """
    params = []
    if filter_employee_ids:
        placeholders = ','.join(['?'] * len(filter_employee_ids))
        query += f" WHERE e.id IN ({placeholders})"
        params.extend(filter_employee_ids)
        
    query += " ORDER BY e.full_name"
    
    employees_list = db.execute(query, params).fetchall()

    # Exportar a Excel (CSV)
    if request.args.get('export') == 'true':
        wb = Workbook()
        ws = wb.active
        ws.title = "Empleados"
        ws.append(['ID', 'Usuario', 'Nombre Completo', 'Email', 'Puesto', 'Departamento', 'Empresa', 'Fecha Contratación', 'Rol', 'Jefe Directo', 'Activo', 'Gestionado por AD'])
        
        for emp in employees_list:
            ws.append([emp['id'], emp['username'], emp['full_name'], emp['email'], emp['job_title'], emp['department'], emp['company'], emp['hire_date'], emp['role'], emp['manager_name'], 'Sí' if emp['is_active'] else 'No', 'Sí' if emp['is_ad_managed'] else 'No'])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": "attachment; filename=lista_empleados.xlsx"}
        )

    return render_template("hr/hr_employee_list.html", employees=employees_list, all_employees=employees, filters={'employee_id': filter_employee_ids})

@bp.route("/employee/add", methods=['GET', 'POST'])
def hr_add_employee():
    if not check_hr_access(readonly=True): # Asistente puede cargar empleados
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    if request.method == 'POST':
        try:
            username = request.form['username']
            full_name = request.form['full_name']
            email = request.form['email']
            password = request.form['password']
            hire_date = datetime.strptime(request.form['hire_date'], '%d/%m/%Y').date()
            role = request.form['role']
            department = request.form['department']
            job_title = request.form['job_title']
            company = request.form['company']
            manager_id = request.form.get('manager_id') or None
        except (ValueError, TypeError):
            flash("Formato de fecha de contratación inválido. Por favor, usa DD/MM/YYYY.", "danger")
            return redirect(url_for('hr.hr_add_employee'))

        db = get_db()
        try:
            db.execute(
                """
                INSERT INTO employees (username, full_name, email, password, hire_date, role, manager_id, department, job_title, company, is_ad_managed)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)
                """,
                (username, full_name, email, generate_password_hash(password), hire_date, role, manager_id, department, job_title, company)
            )
            db.commit()
            flash(f"Empleado '{full_name}' creado exitosamente.", "success")
        except sqlite3.IntegrityError:
            flash(f"El nombre de usuario '{username}' ya existe.", "danger")
        return redirect(url_for('hr.hr_employee_list'))

    db = get_db()
    managers = db.execute("SELECT id, full_name FROM employees WHERE role = 'Jefe' AND is_active = 1").fetchall()
    roles = db.execute("SELECT name FROM roles ORDER BY name").fetchall()
    return render_template("hr/hr_employee_form.html", managers=managers, roles=roles, form_title="Añadir Nuevo Empleado")

@bp.route("/employee/edit/<int:employee_id>", methods=['GET', 'POST'])
def hr_edit_employee(employee_id):
    if not check_hr_access(readonly=True): # Asistente puede editar empleados
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))
    
    db = get_db()
    employee = db.execute("SELECT * FROM employees WHERE id = ?", (employee_id,)).fetchone()
    
    if not employee:
        flash("Empleado no encontrado.", "danger")
        return redirect(url_for('hr.hr_employee_list'))

    if request.method == 'POST':
        password = request.form['password']
        role = request.form['role']
        manager_id = request.form.get('manager_id') or None
        is_active = 'is_active' in request.form

        if employee['is_ad_managed']:
            if password:
                db.execute(
                    "UPDATE employees SET password = ?, role = ?, manager_id = ?, is_active = ? WHERE id = ?",
                    (generate_password_hash(password), role, manager_id, is_active, employee_id)
                )
            else:
                db.execute(
                    "UPDATE employees SET role = ?, manager_id = ?, is_active = ? WHERE id = ?",
                    (role, manager_id, is_active, employee_id)
                )
        else:
            try:
                username = request.form['username']
                full_name = request.form['full_name']
                email = request.form['email']
                hire_date = datetime.strptime(request.form['hire_date'], '%d/%m/%Y').date()
                department = request.form['department']
                job_title = request.form['job_title']
                company = request.form['company']
            except (ValueError, TypeError):
                flash("Formato de fecha de contratación inválido. Por favor, usa DD/MM/YYYY.", "danger")
                return redirect(url_for('hr.hr_edit_employee', employee_id=employee_id))

            try:
                if password:
                    db.execute(
                        """
                        UPDATE employees SET username=?, full_name=?, email=?, password=?, hire_date=?, role=?, manager_id=?, department=?, job_title=?, company=?, is_active=?
                        WHERE id = ?
                        """,
                        (username, full_name, email, generate_password_hash(password), hire_date, role, manager_id, department, job_title, company, is_active, employee_id)
                    )
                else:
                    db.execute(
                        """
                        UPDATE employees SET username=?, full_name=?, email=?, hire_date=?, role=?, manager_id=?, department=?, job_title=?, company=?, is_active=?
                        WHERE id = ?
                        """,
                        (username, full_name, email, hire_date, role, manager_id, department, job_title, company, is_active, employee_id)
                    )
                db.commit()
                flash(f"Empleado '{full_name}' actualizado exitosamente.", "success")
            except sqlite3.IntegrityError:
                flash(f"El nombre de usuario '{username}' ya está en uso por otro empleado.", "danger")
            return redirect(url_for('hr.hr_employee_list'))

        db.commit()
        return redirect(url_for('hr.hr_employee_list'))

    managers = db.execute("SELECT id, full_name FROM employees WHERE role = 'Jefe' AND is_active = 1 AND id != ?", (employee_id,)).fetchall()
    roles = db.execute("SELECT name FROM roles ORDER BY name").fetchall()
    return render_template("hr/hr_employee_form.html", employee=employee, managers=managers, roles=roles, form_title="Editar Empleado")

@bp.route("/request/create", methods=['GET', 'POST'])
def hr_create_request():
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    db = get_db()

    if request.method == 'POST':
        try:
            employee_id = request.form.get('employee_id')
            leave_type_id = request.form.get('leave_type_id')
            start_date_str = request.form.get('start_date')
            replacement_id = request.form.get('replacement_id')
            attachment = request.files.get('attachment')
            
            # Obtener nombre del reemplazo
            replacement_name = None
            if replacement_id:
                rep_emp = db.execute("SELECT full_name FROM employees WHERE id = ?", (replacement_id,)).fetchone()
                if rep_emp:
                    replacement_name = rep_emp['full_name']
            
            if not replacement_name:
                flash("El campo de reemplazo es obligatorio.", "danger")
                return redirect(url_for('hr.hr_create_request'))
            
            # Obtener info del tipo de licencia
            leave_type = db.execute("SELECT * FROM leave_types WHERE id = ?", (leave_type_id,)).fetchone()
            
            start_date = datetime.strptime(start_date_str, "%d/%m/%Y").date()
            
            if leave_type['consumption_type'] == 'Fixed':
                # Días corridos: Calcular fecha fin automáticamente
                days_requested = leave_type['default_days']
                # Restamos 1 porque el día de inicio cuenta
                end_date = start_date + timedelta(days=days_requested - 1)
            else:
                # Días hábiles: Usar fecha fin del formulario
                end_date_str = request.form.get('end_date')
                end_date = datetime.strptime(end_date_str, "%d/%m/%Y").date()
                days_requested = calculate_working_days(start_date, end_date)

            if end_date < start_date:
                flash("La fecha de fin no puede ser anterior a la de inicio.", "danger")
                return redirect(url_for('hr.hr_create_request'))

            # Verificar saldo si es necesario
            if leave_type['requires_balance']:
                balance_row = db.execute(
                    "SELECT SUM(total_days_accrued - days_taken) as balance FROM vacation_periods WHERE employee_id = ? AND leave_type_id = ?",
                    (employee_id, leave_type_id)
                ).fetchone()
                balance = balance_row['balance'] if balance_row and balance_row['balance'] else 0
                
                if balance < days_requested:
                    flash(f"El empleado no tiene saldo suficiente. Saldo: {balance}, Solicitado: {days_requested}.", "danger")
                    return redirect(url_for('hr.hr_create_request'))

            # Manejo de Adjunto
            attachment_path = None
            if leave_type['requires_attachment']:
                if attachment and attachment.filename != '':
                    filename = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{attachment.filename}")
                    attachment.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
                    attachment_path = filename
                else:
                    flash("Este tipo de licencia requiere un archivo adjunto obligatorio.", "danger")
                    return redirect(url_for('hr.hr_create_request'))

            # Insertar solicitud con estado 'Pendiente' pero con hr_approval_date ya seteado
            db.execute(
                """
                INSERT INTO vacation_requests (employee_id, leave_type_id, start_date, end_date, request_type, days_requested, replacement_name, status, hr_approval_date, attachment_path)
                VALUES (?, ?, ?, ?, 'FullDay', ?, ?, 'Pendiente', ?, ?)
                """,
                (employee_id, leave_type_id, start_date, end_date, days_requested, replacement_name, datetime.now(), attachment_path)
            )
            db.commit()

            # Notificar al Jefe
            emp_info = db.execute("SELECT full_name, manager_id FROM employees WHERE id = ?", (employee_id,)).fetchone()
            if emp_info['manager_id']:
                manager = db.execute("SELECT email, full_name FROM employees WHERE id = ?", (emp_info['manager_id'],)).fetchone()
                if manager and manager['email']:
                    subject = f"Solicitud Creada por RRHH: {emp_info['full_name']}"
                    body = f"Estimado/a {manager['full_name']},\n\nRRHH ha cargado una solicitud de licencia/vacaciones para {emp_info['full_name']}.\n\nTipo: {leave_type['name']}\nInicio: {start_date.strftime('%d/%m/%Y')}\nFin: {end_date.strftime('%d/%m/%Y')}\n\nPor favor ingrese al sistema para aprobarla."
                    send_email(subject, [manager['email']], body)

            flash("Solicitud creada exitosamente. Se ha notificado al jefe para su aprobación.", "success")
            return redirect(url_for('hr.hr_all_requests'))

        except (ValueError, TypeError) as e:
            flash(f"Error en los datos: {e}", "danger")
            return redirect(url_for('hr.hr_create_request'))

    # GET
    employees = db.execute("SELECT id, full_name FROM employees WHERE is_active = 1 ORDER BY full_name").fetchall()
    # Serializar leave_types para usar en JS
    leave_types_rows = db.execute("SELECT * FROM leave_types ORDER BY name").fetchall()
    leave_types_json = {}
    for lt in leave_types_rows:
        leave_types_json[lt['id']] = dict(lt)

    # Obtener feriados y sábados para cálculo en JS
    holidays_dict = get_paraguay_holidays()
    holidays_list = [d.strftime('%d/%m/%Y') for d in holidays_dict.keys()]
    working_saturdays_rows = db.execute("SELECT effective_date FROM saturday_config WHERE is_working = 1").fetchall()
    working_saturdays = [row['effective_date'].strftime('%d/%m/%Y') for row in working_saturdays_rows]

    selected_employee_id = request.args.get('employee_id')

    return render_template("hr/hr_create_request.html", employees=employees, leave_types=leave_types_rows, leave_types_json=json.dumps(leave_types_json), selected_employee_id=selected_employee_id, holidays_list=holidays_list, working_saturdays=working_saturdays)

@bp.route('/api/balances/<int:employee_id>')
def get_employee_balances_api(employee_id):
    if not check_hr_access(readonly=True):
        return jsonify({"error": "Unauthorized"}), 403
    
    db = get_db()
    balances = db.execute(
        "SELECT leave_type_id, SUM(total_days_accrued - days_taken) as balance FROM vacation_periods WHERE employee_id = ? GROUP BY leave_type_id",
        (employee_id,)
    ).fetchall()
    
    result = {row['leave_type_id']: row['balance'] for row in balances}
    return jsonify(result)

@bp.route("/all_requests")
def hr_all_requests():
    if not check_hr_access(readonly=True):
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    db = get_db()
    
    # Actualizar estados de solicitudes
    today = date.today()
    # 1. Aprobado por RRHH -> Activo
    db.execute(
        "UPDATE vacation_requests SET status = 'Activo' WHERE status = 'Aprobado por RRHH' AND start_date <= ? AND end_date >= ?",
        (today, today)
    )
    # 2. Activo/Aprobado -> Finalizado
    db.execute(
        "UPDATE vacation_requests SET status = 'Finalizado' WHERE (status = 'Aprobado por RRHH' OR status = 'Activo') AND end_date < ?",
        (today,)
    )
    db.commit()
    
    filter_employee_ids = request.args.getlist('employee_id')
    filter_status = request.args.get('status', '')
    filter_type = request.args.get('type', '')
    filter_date_from = request.args.get('date_from', '')
    filter_date_to = request.args.get('date_to', '')

    query = """
        SELECT vr.*, e.full_name as employee_name, m.full_name as manager_name, lt.name as leave_name
        FROM vacation_requests vr
        JOIN employees e ON vr.employee_id = e.id
        LEFT JOIN leave_types lt ON vr.leave_type_id = lt.id
        LEFT JOIN employees m ON e.manager_id = m.id
    """
    conditions = []
    params = []

    if filter_employee_ids:
        placeholders = ','.join(['?'] * len(filter_employee_ids))
        conditions.append(f"vr.employee_id IN ({placeholders})")
        params.extend(filter_employee_ids)
    
    if filter_status:
        conditions.append("vr.status = ?")
        params.append(filter_status)
        
    if filter_type:
        conditions.append("vr.request_type = ?")
        params.append(filter_type)

    if filter_date_from:
        try:
            d_from = datetime.strptime(filter_date_from, '%d/%m/%Y').date()
            conditions.append("vr.start_date >= ?")
            params.append(d_from)
        except ValueError:
            pass

    if filter_date_to:
        try:
            d_to = datetime.strptime(filter_date_to, '%d/%m/%Y').date()
            conditions.append("vr.start_date <= ?")
            params.append(d_to)
        except ValueError:
            pass

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " ORDER BY vr.request_date DESC"

    all_requests = db.execute(query, params).fetchall()

    # Exportar a Excel (CSV)
    if request.args.get('export') == 'true':
        wb = Workbook()
        ws = wb.active
        ws.title = "Todas las Solicitudes"
        ws.append(['ID', 'Empleado', 'Jefe Directo', 'Tipo Licencia', 'Tipo Solicitud', 'Inicio', 'Fin', 'Días', 'Estado', 'Fecha Solicitud'])
        
        for req in all_requests:
            ws.append([req['id'], req['employee_name'], req['manager_name'], req['leave_name'], req['request_type'], req['start_date'], req['end_date'], req['days_requested'], req['status'], req['request_date']])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": "attachment; filename=todas_las_solicitudes.xlsx"}
        )
    
    employees = db.execute("SELECT id, full_name FROM employees ORDER BY full_name").fetchall()
    
    return render_template("hr/hr_all_requests.html", 
                           requests=all_requests, 
                           employees=employees, 
                           filters={
                               'employee_id': filter_employee_ids, 
                               'status': filter_status, 
                               'type': filter_type,
                               'date_from': filter_date_from,
                               'date_to': filter_date_to
                           })

@bp.route("/holidays", methods=['GET', 'POST'])
def hr_manage_holidays():
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    db = get_db()

    if request.method == 'POST':
        try:
            holiday_date_str = request.form['holiday_date']
            description = request.form['description']
            holiday_date = datetime.strptime(holiday_date_str, '%d/%m/%Y').date()
            
            # Capturamos el checkbox. Si está marcado devuelve algo (True), si no, None (False)
            is_recurring = 1 if request.form.get('is_recurring') else 0

            # Si es recurrente y la fecha ya pasó, actualizar al próximo año válido
            if is_recurring:
                today = datetime.now().date()
                if holiday_date < today:
                    try:
                        # Intentar con el año actual
                        candidate_date = holiday_date.replace(year=today.year)
                    except ValueError:
                        # Manejo de 29 de febrero en años no bisiestos
                        candidate_date = datetime(today.year, 3, 1).date()
                    
                    if candidate_date < today:
                        try:
                            # Si ya pasó este año, mover al siguiente
                            candidate_date = holiday_date.replace(year=today.year + 1)
                        except ValueError:
                            candidate_date = datetime(today.year + 1, 3, 1).date()
                    
                    holiday_date = candidate_date
                    flash(f"Fecha actualizada automáticamente a la próxima ocurrencia: {holiday_date.strftime('%d/%m/%Y')}", "info")
        except (ValueError, TypeError):
            flash("Formato de fecha inválido. Por favor, usa DD/MM/YYYY.", "danger")
            return redirect(url_for('hr.hr_manage_holidays'))
        
        if not holiday_date or not description:
            flash("Ambos campos, fecha y descripción, son obligatorios.", "danger")
        else:
            try:
                db.execute(
                    "INSERT INTO custom_holidays (holiday_date, description, is_recurring) VALUES (?, ?, ?)",
                    (holiday_date, description, is_recurring)
                )
                db.commit()
                flash("Feriado añadido exitosamente.", "success")
            except sqlite3.IntegrityError:
                flash(f"La fecha {holiday_date.strftime('%d/%m/%Y')} ya está registrada como feriado.", "warning")
        
        return redirect(url_for('hr.hr_manage_holidays'))

    # Actualizar automáticamente feriados recurrentes vencidos al visualizar la lista
    today = datetime.now().date()
    recurring_holidays = db.execute("SELECT id, holiday_date FROM custom_holidays WHERE is_recurring = 1").fetchall()
    updates_made = False

    for hol in recurring_holidays:
        h_date = hol['holiday_date']
        if h_date < today:
            try:
                # Intentar actualizar al año actual
                next_date = h_date.replace(year=today.year)
            except ValueError:
                next_date = datetime(today.year, 3, 1).date()
            
            if next_date < today:
                try:
                    next_date = h_date.replace(year=today.year + 1)
                except ValueError:
                    next_date = datetime(today.year + 1, 3, 1).date()
            
            # Verificar que la nueva fecha no exista ya (para evitar error UNIQUE)
            if not db.execute("SELECT id FROM custom_holidays WHERE holiday_date = ? AND id != ?", (next_date, hol['id'])).fetchone():
                db.execute("UPDATE custom_holidays SET holiday_date = ? WHERE id = ?", (next_date, hol['id']))
                updates_made = True

    if updates_made:
        db.commit()

    holidays_list = db.execute("SELECT id, holiday_date, description, is_recurring FROM custom_holidays ORDER BY holiday_date DESC").fetchall()
    # Cargar también la configuración de sábados para la nueva pestaña
    saturdays_list = db.execute("SELECT * FROM saturday_config ORDER BY effective_date ASC").fetchall()
    
    return render_template("hr/hr_holidays.html", holidays=holidays_list, saturdays=saturdays_list)

@bp.route("/holiday/delete/<int:holiday_id>", methods=['POST'])
def hr_delete_holiday(holiday_id):
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    db = get_db()
    db.execute("DELETE FROM custom_holidays WHERE id = ?", (holiday_id,))
    db.commit()
    
    flash("Feriado eliminado exitosamente.", "success")
    return redirect(url_for('hr.hr_manage_holidays'))

@bp.route('/ad_sync', methods=['GET', 'POST'])
def hr_ad_sync():
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    config_path = current_app.config['AD_CONFIG_PATH']
    
    if request.method == 'POST':
        config = {
            'server': request.form['server'],
            'port': int(request.form['port']),
            'use_ssl': 'use_ssl' in request.form,
            'user': request.form['user'],
            'password': request.form['password'],
            'search_base': request.form['search_base'],
            'email_attribute': request.form['email_attribute'],
            'department_attribute': request.form['department_attribute'],
            'hire_date_attribute': request.form['hire_date_attribute'],
            'hire_date_format': request.form['hire_date_format'],
            'job_title_attribute': request.form['job_title_attribute'],
            'company_attribute': request.form['company_attribute'],
        }
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=4)
        flash("Configuración de Directorio Activo guardada.", "success")
        return redirect(url_for('hr.hr_ad_sync'))

    config = {}
    if os.path.exists(config_path):
        with open(config_path, 'r') as f:
            config = json.load(f)
            
    return render_template('hr/hr_ad_sync.html', config=config)

@bp.route("/roles", methods=['GET', 'POST'])
def hr_manage_roles():
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))
    
    db = get_db()
    
    if request.method == 'POST':
        role_name = request.form.get('name')
        base_role = request.form.get('base_role')
        if role_name and base_role:
            try:
                db.execute("INSERT INTO roles (name, base_role, is_system_role) VALUES (?, ?, 0)", (role_name, base_role))
                db.commit()
                flash(f"Rol '{role_name}' creado.", "success")
            except sqlite3.IntegrityError:
                flash("El rol ya existe.", "danger")
        return redirect(url_for('hr.hr_manage_roles'))

    roles = db.execute("SELECT * FROM roles ORDER BY name").fetchall()
    return render_template("hr/hr_roles.html", roles=roles)

@bp.route("/roles/delete/<int:role_id>", methods=['POST'])
def hr_delete_role(role_id):
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))
        
    db = get_db()
    role = db.execute("SELECT * FROM roles WHERE id = ?", (role_id,)).fetchone()
    
    if role:
        if role['is_system_role']:
            flash("No se pueden eliminar roles del sistema.", "danger")
        else:
            db.execute("DELETE FROM roles WHERE id = ?", (role_id,))
            db.commit()
            flash("Rol eliminado.", "success")
    
    return redirect(url_for('hr.hr_manage_roles'))


@bp.route('/cancellations')
def hr_cancellation_list():
    if not check_hr_access(readonly=True):
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    db = get_db()
    
    filter_employee_ids = request.args.getlist('employee_id')
    
    query = """
        SELECT vr.id, vr.start_date, vr.end_date, vr.days_requested, e.full_name as employee_name
        FROM vacation_requests vr
        JOIN employees e ON vr.employee_id = e.id
        WHERE vr.status = 'Anulación Pendiente RRHH'
    """
    params = []
    if filter_employee_ids:
        placeholders = ','.join(['?'] * len(filter_employee_ids))
        query += f" AND vr.employee_id IN ({placeholders})"
        params.extend(filter_employee_ids)
        
    query += " ORDER BY vr.request_date"
    
    cancellation_requests = db.execute(query, params).fetchall()

    # Exportar a Excel (CSV)
    if request.args.get('export') == 'true':
        wb = Workbook()
        ws = wb.active
        ws.title = "Anulaciones"
        ws.append(['ID', 'Empleado', 'Inicio', 'Fin', 'Días a Devolver'])
        
        for req in cancellation_requests:
            ws.append([req['id'], req['employee_name'], req['start_date'], req['end_date'], req['days_requested']])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": "attachment; filename=solicitudes_anulacion.xlsx"}
        )

    employees = db.execute("SELECT id, full_name FROM employees ORDER BY full_name").fetchall()
    
    return render_template("hr/hr_cancellation_list.html", requests=cancellation_requests, employees=employees, filters={'employee_id': filter_employee_ids})

@bp.route('/cancellation/approve/<int:request_id>', methods=['POST'])
def hr_approve_cancellation(request_id):
    if not check_hr_access(): # Solo RRHH
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    db = get_db()
    req = db.execute(
        "SELECT employee_id, days_requested, leave_type_id, replacement_name FROM vacation_requests WHERE id = ? AND status = 'Anulación Pendiente RRHH'",
        (request_id,)
    ).fetchone()

    if req:
        leave_type = db.execute("SELECT requires_balance FROM leave_types WHERE id = ?", (req['leave_type_id'],)).fetchone()
        requires_balance = leave_type['requires_balance'] if leave_type else 1

        if requires_balance:
            days_to_refund = req['days_requested']
            
            # Obtener los periodos del empleado ordenados del más nuevo al más antiguo
            periods = db.execute(
                "SELECT id, days_taken FROM vacation_periods WHERE employee_id = ? ORDER BY year DESC",
                (req['employee_id'],)
            ).fetchall()

            for period in periods:
                if days_to_refund <= 0:
                    break
                
                refundable_from_this_period = min(days_to_refund, period['days_taken'])
                
                if refundable_from_this_period > 0:
                    db.execute(
                        "UPDATE vacation_periods SET days_taken = days_taken - ? WHERE id = ?",
                        (refundable_from_this_period, period['id'])
                    )
                    days_to_refund -= refundable_from_this_period

        db.execute(
            "UPDATE vacation_requests SET status = 'Anulado' WHERE id = ?",
            (request_id,)
        )
        db.commit()

        # --- NOTIFICACIÓN: A Empleado, Jefe y Reemplazo (Anulación Aprobada) ---
        try:
            emp_info = db.execute("SELECT email, full_name, manager_id FROM employees WHERE id = ?", (req['employee_id'],)).fetchone()
            recipients = []

            if emp_info and emp_info['email']:
                recipients.append(emp_info['email'])
            
            if emp_info and emp_info['manager_id']:
                mgr_info = db.execute("SELECT email FROM employees WHERE id = ?", (emp_info['manager_id'],)).fetchone()
                if mgr_info and mgr_info['email']:
                    recipients.append(mgr_info['email'])

            if req['replacement_name']:
                rep_info = db.execute("SELECT email FROM employees WHERE full_name = ?", (req['replacement_name'],)).fetchone()
                if rep_info and rep_info['email']:
                    recipients.append(rep_info['email'])

            # Obtener email de RRHH (Actor) para CC
            hr_actor_email = db.execute("SELECT email FROM employees WHERE id = ?", (session['user_id'],)).fetchone()
            cc_list = [hr_actor_email['email']] if hr_actor_email and hr_actor_email['email'] else []

            if recipients:
                subject = "Anulación de Vacaciones Confirmada"
                body = f"Estimado/a,\n\nLa solicitud de anulación de vacaciones de {emp_info['full_name']} ha sido procesada y los días han sido devueltos al saldo.\nLa vacación queda cancelada."
                send_email(subject, recipients, body, cc=cc_list)
        except Exception as e:
            print(f"Error enviando email al empleado: {e}")

        flash('Anulación aprobada y días devueltos al saldo del empleado.', 'success')
    else:
        flash('No se pudo encontrar la solicitud o ya fue procesada.', 'warning')

    return redirect(url_for('hr.hr_cancellation_list'))

@bp.route('/cancellation/reject/<int:request_id>', methods=['POST'])
def hr_reject_cancellation(request_id):
    if not check_hr_access(): # Solo RRHH
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))
    
    db = get_db()
    req = db.execute(
        "SELECT id FROM vacation_requests WHERE id = ? AND status = 'Anulación Pendiente RRHH'",
        (request_id,)
    ).fetchone()

    if req:
        db.execute(
            "UPDATE vacation_requests SET status = 'Aprobado por RRHH' WHERE id = ?",
            (request_id,)
        )
        db.commit()

        # --- NOTIFICACIÓN: A Empleado y Jefe (Anulación Rechazada) ---
        try:
            emp_info = db.execute("SELECT e.email, e.full_name, e.manager_id FROM vacation_requests vr JOIN employees e ON vr.employee_id = e.id WHERE vr.id = ?", (request_id,)).fetchone()
            recipients = []

            if emp_info and emp_info['email']:
                recipients.append(emp_info['email'])
            
            if emp_info and emp_info['manager_id']:
                mgr_info = db.execute("SELECT email FROM employees WHERE id = ?", (emp_info['manager_id'],)).fetchone()
                if mgr_info and mgr_info['email']:
                    recipients.append(mgr_info['email'])

            # Obtener email de RRHH (Actor) para CC
            hr_actor_email = db.execute("SELECT email FROM employees WHERE id = ?", (session['user_id'],)).fetchone()
            cc_list = [hr_actor_email['email']] if hr_actor_email and hr_actor_email['email'] else []

            if recipients:
                subject = "Solicitud de Anulación Rechazada"
                body = f"Estimado/a,\n\nLa solicitud de anulación de vacaciones de {emp_info['full_name']} ha sido rechazada por RRHH. La vacación original sigue vigente."
                send_email(subject, recipients, body, cc=cc_list)
        except Exception as e:
            print(f"Error enviando email al empleado: {e}")

        flash('Se ha rechazado la solicitud de anulación.', 'info')
    else:
        flash('No se pudo encontrar la solicitud o ya fue procesada.', 'warning')

    return redirect(url_for('hr.hr_cancellation_list'))


@bp.route('/trigger_ad_sync', methods=['POST'])
def trigger_ad_sync():
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    config_path = current_app.config['AD_CONFIG_PATH']
    db_path = current_app.config['DATABASE']
    
    config = {}
    if os.path.exists(config_path):
        with open(config_path, 'r') as f:
            config = json.load(f)

    if not config:
        flash("La configuración de Directorio Activo no ha sido establecida.", "danger")
        return redirect(url_for('hr.hr_ad_sync'))

    try:
        summary = ad_sync.sync_users_from_ad(config)
        flash(f"Sincronización completada. {summary['created']} usuarios creados, {summary['updated']} actualizados, {summary['deactivated']} desactivados.", "success")
    except Exception as e:
        flash(f"Error durante la sincronización: {e}", "danger")

    return redirect(url_for('hr.hr_ad_sync'))


@bp.route('/team_calendar')
def team_calendar():
    if session.get("role") not in ["Jefe", "RRHH", "Asistente RRHH", "Empleado"]:
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    db = get_db()
    
    approved_requests = db.execute(
        """
        SELECT e.full_name, vr.start_date, vr.end_date, lt.name as leave_name
        FROM vacation_requests vr
        JOIN employees e ON vr.employee_id = e.id
        LEFT JOIN leave_types lt ON vr.leave_type_id = lt.id
        WHERE vr.status IN ('Aprobado por RRHH', 'Activo', 'Finalizado')
        """
    ).fetchall()

    events = []
    for req in approved_requests:
        start_date_obj = req['start_date']
        end_date_obj = req['end_date'] + timedelta(days=1)
        events.append({
            'title': f"{req['full_name']} ({req['leave_name'] or 'Vacaciones'})",
            'start': start_date_obj.strftime('%Y-%m-%d'),
            'end': end_date_obj.strftime('%Y-%m-%d'),
            'allDay': True
        })

    # Agregar Sábados LIBRES (Feriados) al calendario
    saturdays = db.execute("SELECT effective_date FROM saturday_config WHERE is_working = 0").fetchall()
    for sat in saturdays:
        events.append({
            'title': 'Sábado Libre',
            'start': sat['effective_date'].strftime('%Y-%m-%d'),
            'allDay': True,
            'display': 'background', # Muestra como fondo coloreado
            'backgroundColor': '#ffc107' # Color amarillo/ámbar
        })

    # Agregar Feriados (Nacionales y Personalizados)
    holidays = get_paraguay_holidays()
    for date_obj, name in holidays.items():
        events.append({
            'title': name,
            'start': date_obj.strftime('%Y-%m-%d'),
            'allDay': True,
            'display': 'background',
            'backgroundColor': '#ffc107'
        })
        
    return render_template('team_calendar.html', events=json.dumps(events))

@bp.route("/saturdays", methods=['POST'])
def hr_manage_saturdays():
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))
        
    db = get_db()
    action = request.form.get('action')

    if action == 'generate':
        date_str = request.form.get('start_date')
        try:
            start_date = datetime.strptime(date_str, '%d/%m/%Y').date()
            if start_date.weekday() != 5:
                flash("La fecha seleccionada debe ser un sábado.", "danger")
            else:
                # Generar intercalado hasta fin de año
                current_year = start_date.year
                curr = start_date
                working = True # El seleccionado es laboral
                count = 0
                
                while curr.year == current_year:
                    # Upsert: Actualizar si existe, insertar si no
                    exists = db.execute("SELECT id FROM saturday_config WHERE effective_date = ?", (curr,)).fetchone()
                    if exists:
                        db.execute("UPDATE saturday_config SET is_working = ? WHERE id = ?", (1 if working else 0, exists['id']))
                    else:
                        db.execute("INSERT INTO saturday_config (effective_date, is_working) VALUES (?, ?)", (curr, 1 if working else 0))
                    
                    curr += timedelta(days=7)
                    working = not working # Intercalar
                    count += 1

                db.commit()
                flash(f"Se generaron {count} sábados hasta fin de año comenzando el {start_date.strftime('%d/%m/%Y')}.", "success")
        except ValueError:
            flash("Fecha inválida.", "danger")
            
    elif action == 'delete_all':
        db.execute("DELETE FROM saturday_config")
        db.commit()
        flash("Se han eliminado todos los registros de sábados. Ahora todos los sábados se consideran libres por defecto.", "info")

    return redirect(url_for('hr.hr_manage_holidays'))

@bp.route("/leave_types", methods=['GET', 'POST'])
def hr_manage_leave_types():
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))
        
    db = get_db()
    if request.method == 'POST':
        name = request.form.get('name')
        requires_balance = 1 if request.form.get('requires_balance') else 0
        default_days = request.form.get('default_days', 0)
        consumption_type = request.form.get('consumption_type', 'Flexible')
        requires_attachment = 1 if request.form.get('requires_attachment') else 0

        if name:
            try:
                db.execute("INSERT INTO leave_types (name, requires_balance, default_days, consumption_type, requires_attachment) VALUES (?, ?, ?, ?, ?)", (name, requires_balance, default_days, consumption_type, requires_attachment))
                db.commit()
                flash("Tipo de día libre creado.", "success")
            except sqlite3.IntegrityError:
                flash("El nombre ya existe.", "danger")
        return redirect(url_for('hr.hr_manage_leave_types'))
        
    types = db.execute("SELECT * FROM leave_types").fetchall()
    return render_template("hr/hr_leave_types.html", types=types)

@bp.route("/leave_types/delete/<int:type_id>", methods=['POST'])
def hr_delete_leave_type(type_id):
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    db = get_db()
    leave_type = db.execute("SELECT name FROM leave_types WHERE id = ?", (type_id,)).fetchone()
    
    if leave_type:
        if leave_type['name'] == 'Vacaciones':
            flash("El tipo de licencia 'Vacaciones' es del sistema y no se puede eliminar.", "danger")
        else:
            db.execute("DELETE FROM leave_types WHERE id = ?", (type_id,))
            db.commit()
            flash("Tipo de licencia eliminado.", "success")
    
    return redirect(url_for('hr.hr_manage_leave_types'))

@bp.route("/interrupt_vacation/<int:request_id>", methods=['POST'])
def hr_interrupt_vacation(request_id):
    if not check_hr_access():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))
        
    db = get_db()
    reintegration_date_str = request.form.get('reintegration_date')
    interruption_reason = request.form.get('interruption_reason')
    
    if not reintegration_date_str or not interruption_reason:
        flash("La fecha de reintegración y el comentario son obligatorios.", "danger")
        return redirect(url_for('hr.hr_all_requests'))

    try:
        reintegration_date = datetime.strptime(reintegration_date_str, '%d/%m/%Y').date()
    except ValueError:
        flash("Fecha inválida. Use el formato DD/MM/YYYY.", "danger")
        return redirect(url_for('hr.hr_all_requests'))
        
    req = db.execute("SELECT * FROM vacation_requests WHERE id = ?", (request_id,)).fetchone()
    if not req or req['status'] not in ['Aprobado por RRHH', 'Activo']:
        flash("Solicitud no válida para interrupción.", "danger")
        return redirect(url_for('hr.hr_all_requests'))
    
    if reintegration_date <= req['start_date']:
        flash("La fecha de reintegración debe ser posterior a la fecha de inicio.", "danger")
        return redirect(url_for('hr.hr_all_requests'))

    # La vacación termina el día anterior a la reintegración
    new_end_date = reintegration_date - timedelta(days=1)
        
    from ..utils import calculate_working_days
    days_used = calculate_working_days(req['start_date'], new_end_date)
    days_refund = req['days_requested'] - days_used
    
    if days_refund > 0:
        # Devolver días a los periodos correspondientes (LIFO)
        remaining_refund = days_refund
        periods = db.execute(
            "SELECT id, days_taken FROM vacation_periods WHERE employee_id = ? AND days_taken > 0 ORDER BY year DESC", 
            (req['employee_id'],)
        ).fetchall()
        
        for period in periods:
            if remaining_refund <= 0:
                break
            
            refund_amount = min(remaining_refund, period['days_taken'])
            db.execute("UPDATE vacation_periods SET days_taken = days_taken - ? WHERE id = ?", (refund_amount, period['id']))
            remaining_refund -= refund_amount

    db.execute("UPDATE vacation_requests SET end_date = ?, days_requested = ?, interruption_reason = ? WHERE id = ?", (new_end_date, days_used, interruption_reason, request_id))
    db.commit()
    flash(f"Vacación interrumpida. Fecha de fin actualizada a {new_end_date.strftime('%d/%m/%Y')}. Se devolvieron {days_refund} días al saldo.", "success")

    # --- NOTIFICACIÓN: Corte de Vacaciones (Interrupción) ---
    try:
        emp_info = db.execute("SELECT email, full_name, manager_id FROM employees WHERE id = ?", (req['employee_id'],)).fetchone()
        recipients = []

        if emp_info and emp_info['email']:
            recipients.append(emp_info['email'])
        
        if emp_info and emp_info['manager_id']:
            mgr_info = db.execute("SELECT email FROM employees WHERE id = ?", (emp_info['manager_id'],)).fetchone()
            if mgr_info and mgr_info['email']:
                recipients.append(mgr_info['email'])

        if req['replacement_name']:
            rep_info = db.execute("SELECT email FROM employees WHERE full_name = ?", (req['replacement_name'],)).fetchone()
            if rep_info and rep_info['email']:
                recipients.append(rep_info['email'])

        # Obtener email de RRHH (Actor) para CC
        hr_actor_email = db.execute("SELECT email FROM employees WHERE id = ?", (session['user_id'],)).fetchone()
        cc_list = [hr_actor_email['email']] if hr_actor_email and hr_actor_email['email'] else []

        if recipients:
            subject = "Notificación de Corte de Vacaciones"
            body = f"Estimado/a,\n\nSe ha realizado un corte en las vacaciones de {emp_info['full_name']}.\n\n- Nueva fecha de fin: {new_end_date.strftime('%d/%m/%Y')}\n- Motivo del corte: {interruption_reason}\n\nLos días restantes han sido devueltos al saldo."
            send_email(subject, recipients, body, cc=cc_list)
    except Exception as e:
        print(f"Error enviando email de interrupción: {e}")
        
    return redirect(url_for('hr.hr_all_requests'))

@bp.route("/modify_days/<int:request_id>", methods=['POST'])
def hr_modify_request_days(request_id):
    # Solo RRHH puede modificar días, Asistente NO.
    if not check_hr_access(readonly=False):
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("main.dashboard"))

    db = get_db()
    try:
        new_start_str = request.form.get('new_start_date')
        new_end_str = request.form.get('new_end_date')
        reason = request.form.get('modification_reason')

        if not new_start_str or not new_end_str:
            flash("Las fechas de inicio y fin son obligatorias.", "danger")
            return redirect(url_for('hr.hr_all_requests'))

        new_start_date = datetime.strptime(new_start_str, '%d/%m/%Y').date()
        new_end_date = datetime.strptime(new_end_str, '%d/%m/%Y').date()
        
        if new_end_date < new_start_date:
            flash("La fecha de fin no puede ser anterior a la de inicio.", "danger")
            return redirect(url_for('hr.hr_all_requests'))
            
        # Recalcular días hábiles según las nuevas fechas
        new_days = calculate_working_days(new_start_date, new_end_date)
        
    except (ValueError, TypeError):
        flash("Formato de fecha inválido.", "danger")
        return redirect(url_for('hr.hr_all_requests'))

    if not reason:
        flash("El motivo de la modificación es obligatorio.", "danger")
        return redirect(url_for('hr.hr_all_requests'))

    req = db.execute("SELECT * FROM vacation_requests WHERE id = ?", (request_id,)).fetchone()
    if not req:
        flash("Solicitud no encontrada.", "danger")
        return redirect(url_for('hr.hr_all_requests'))

    if req['status'] not in ['Aprobado por RRHH', 'Activo']:
        flash("Solo se pueden modificar solicitudes con estado 'Aprobado por RRHH' o 'Activo'.", "danger")
        return redirect(url_for('hr.hr_all_requests'))

    old_days = req['days_requested']
    diff = new_days - old_days

    if diff == 0 and new_start_date == req['start_date'] and new_end_date == req['end_date']:
        flash("No hubo cambios en las fechas ni en la cantidad de días.", "info")
        return redirect(url_for('hr.hr_all_requests'))

    # Si se aumentan los días (diff > 0), hay que descontar más saldo
    if diff > 0:
        # Verificar saldo disponible
        periods = db.execute(
            "SELECT id, total_days_accrued, days_taken FROM vacation_periods WHERE employee_id = ? AND total_days_accrued > days_taken ORDER BY year ASC",
            (req['employee_id'],)
        ).fetchall()
        
        total_balance = sum(p['total_days_accrued'] - p['days_taken'] for p in periods)
        if total_balance < diff:
            flash(f"El empleado no tiene saldo suficiente para agregar {diff} días. Saldo disponible: {total_balance}.", "danger")
            return redirect(url_for('hr.hr_all_requests'))

        # Descontar saldo
        days_to_deduct = diff
        for period in periods:
            if days_to_deduct <= 0: break
            balance = period['total_days_accrued'] - period['days_taken']
            deduct = min(days_to_deduct, balance)
            db.execute("UPDATE vacation_periods SET days_taken = days_taken + ? WHERE id = ?", (deduct, period['id']))
            days_to_deduct -= deduct

    # Si se disminuyen los días (diff < 0), hay que devolver saldo
    elif diff < 0:
        # Devolver saldo (usando lógica LIFO para devoluciones, similar a cancelaciones)
        days_to_refund = abs(diff)
        periods = db.execute(
            "SELECT id, days_taken FROM vacation_periods WHERE employee_id = ? AND days_taken > 0 ORDER BY year DESC",
            (req['employee_id'],)
        ).fetchall()
        
        for period in periods:
            if days_to_refund <= 0: break
            refund = min(days_to_refund, period['days_taken'])
            db.execute("UPDATE vacation_periods SET days_taken = days_taken - ? WHERE id = ?", (refund, period['id']))
            days_to_refund -= refund

    # Actualizar solicitud
    db.execute(
        "UPDATE vacation_requests SET start_date = ?, end_date = ?, days_requested = ?, modification_reason = ? WHERE id = ?",
        (new_start_date, new_end_date, new_days, reason, request_id)
    )
    db.commit()
    
    flash(f"Solicitud modificada exitosamente. Nuevas fechas: {new_start_date.strftime('%d/%m/%Y')} - {new_end_date.strftime('%d/%m/%Y')} ({new_days} días).", "success")

    # --- NOTIFICACIÓN: Reajuste de Días (Modificación) ---
    try:
        emp_info = db.execute("SELECT email, full_name, manager_id FROM employees WHERE id = ?", (req['employee_id'],)).fetchone()
        recipients = []

        if emp_info and emp_info['email']:
            recipients.append(emp_info['email'])
        
        if emp_info and emp_info['manager_id']:
            mgr_info = db.execute("SELECT email FROM employees WHERE id = ?", (emp_info['manager_id'],)).fetchone()
            if mgr_info and mgr_info['email']:
                recipients.append(mgr_info['email'])

        if req['replacement_name']:
            rep_info = db.execute("SELECT email FROM employees WHERE full_name = ?", (req['replacement_name'],)).fetchone()
            if rep_info and rep_info['email']:
                recipients.append(rep_info['email'])

        # Obtener email de RRHH (Actor) para CC
        hr_actor_email = db.execute("SELECT email FROM employees WHERE id = ?", (session['user_id'],)).fetchone()
        cc_list = [hr_actor_email['email']] if hr_actor_email and hr_actor_email['email'] else []

        if recipients:
            subject = "Notificación de Reajuste de Días de Vacaciones"
            body = f"Estimado/a,\n\nSe ha realizado una modificación en la solicitud de vacaciones de {emp_info['full_name']}.\n\n- Nuevas Fechas: {new_start_date.strftime('%d/%m/%Y')} al {new_end_date.strftime('%d/%m/%Y')}\n- Días anteriores: {old_days}\n- Días actuales: {new_days}\n- Motivo: {reason}"
            send_email(subject, recipients, body, cc=cc_list)
    except Exception as e:
        print(f"Error enviando email de modificación: {e}")

    return redirect(url_for('hr.hr_all_requests'))